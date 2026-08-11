"""Excel document loader using openpyxl. Each worksheet is treated as one page."""

import io

import openpyxl

from app.ingestion.base import LoadedDocument, LoadedPage


class ExcelLoader:
    """Loads text from Excel (.xlsx) workbooks, one page per worksheet."""

    def load(self, data: bytes, filename: str) -> LoadedDocument:
        """Render each worksheet's rows as comma-joined text lines."""
        workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        pages = []
        for index, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            lines = [
                ", ".join("" if cell is None else str(cell) for cell in row)
                for row in sheet.iter_rows(values_only=True)
            ]
            pages.append(LoadedPage(text="\n".join(lines), page_number=index + 1))
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return LoadedDocument(source=filename, content_type=content_type, pages=pages)
